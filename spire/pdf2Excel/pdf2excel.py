from spire.pdf.common import *
from spire.pdf import *
import os
import openpyxl

input_pdf = "Sample.pdf"
output_excel = "output.xlsx"

def extract_text_from_pdf(pdf_path):
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")

    doc = PdfDocument()
    doc.LoadFromFile(pdf_path)

    all_pages_text = []
    for i in range(doc.Pages.Count):
        page = doc.Pages[i]
        extractor = PdfTextExtractor(page)
        options = PdfTextExtractOptions()
        text = extractor.ExtractText(options)
        all_pages_text.append(text)

    doc.Close()
    return all_pages_text

def create_excel_file(text_pages, filename):
    wb = openpyxl.Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    for i, page_text in enumerate(text_pages, 1):
        ws = wb.create_sheet(title=f"Page_{i}")
        for row_idx, line in enumerate(page_text.splitlines(), 1):
            ws.cell(row=row_idx, column=1, value=line)

    wb.save(filename)
    print(f"Excel file saved: {filename}")

def main():
    try:
        pages_text = extract_text_from_pdf(input_pdf)
        create_excel_file(pages_text, output_excel)
        print("Excel workbook created successfully!")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
